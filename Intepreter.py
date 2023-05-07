from Actions.CheckTitlePageAction import CheckTitlePageAction
from Actions.CheckPageOrdersAction import CheckPageOrdersAction
from Actions.CheckNavbarAction import CheckNavbarAction
from Actions.CheckPagesEmployees import CheckPageEmployeesAction
from Actions.CheckPagesCustomersAction import CheckPageCustomersAction
from Actions.CheckAddCalendarAction import CheckAddCalendarsAction
from Actions.CheckEditCalendarAction import CheckEditCalendarsAction
from Actions.CheckEditorAction import CheckEditorsAction
from Actions.CheckChartsAction import CheckChartsAction
from Actions.NavigateAction import NavigateAction
import openpyxl
from pathlib import Path
from selenium import webdriver

def main():
    print('Begin')
    xlsx_file = Path('Scripts', 'SampleScripts.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)

    # Read the active sheet:
    sheet = wb_obj.active
    for row in sheet.iter_rows(max_row=sheet.max_row):
        actionName = row[0].value
        print(actionName)
        if actionName == "Test Begin":
            continue
        if actionName == "Test End":
            break
        if actionName == NavigateAction.getActionName():
            valueScript = row[1].value
            NavigateAction(valueScript).doAction()

        if actionName == CheckTitlePageAction.getActionName():
            valueScript = row[1].value
            CheckTitlePageAction(valueScript).doAction()

        if actionName == CheckNavbarAction.getActionName():
            valueScript_cart = row[1].value
            valueScript_arase = row[2].value
            valueScript_chat = row[3].value
            valueScript_noti = row[4].value
            valueScript_profile = row[5].value
            valueScriptId_erase = row[6].value
            CheckNavbarAction(valueScript_cart, valueScript_arase, valueScript_chat, valueScript_noti,
                              valueScript_profile, valueScriptId_erase).doAction()

        if actionName == CheckPageOrdersAction.getActionName():
            valueScript_order = row[1].value
            valueScript_page = row[2].value
            CheckPageOrdersAction(valueScript_order, valueScript_page).doAction()

        if actionName == CheckPageEmployeesAction.getActionName():
            valueScript_employees = row[1].value
            valueScript_page = row[2].value
            CheckPageEmployeesAction(valueScript_employees, valueScript_page).doAction()

        if actionName == CheckPageCustomersAction.getActionName():
            valueScript_customers = row[1].value
            valueScript_item_1 = row[2].value
            valueScript_item_2 = row[3].value
            valueScript_item_3 = row[4].value
            valueScript_delete_1 = row[5].value
            valueScript_delete_all = row[6].value
            valueScript_click = row[7].value
            valueScript_page = row[8].value
            valueScript_select = row[9].value
            valueScript_delete_2 = row[10].value
            CheckPageCustomersAction(valueScript_customers, valueScript_item_1, valueScript_item_2, valueScript_item_3,
                                     valueScript_delete_1, valueScript_delete_all, valueScript_click,
                                     valueScript_page, valueScript_select, valueScript_delete_2).doAction()

        if actionName == CheckAddCalendarsAction.getActionName():
            valueScript_calendar = row[1].value
            valueScriptId_add_task = row[2].value
            valueScript_detail = row[3].value
            valueScriptId_title = row[4].value
            valueScriptId_location = row[5].value
            valueScriptId_description = row[6].value
            valueScript_title = row[7].value
            valueScript_location = row[8].value
            valueScript_description = row[9].value
            value_save_1 = row[10].value
            CheckAddCalendarsAction(valueScript_calendar, valueScriptId_add_task, valueScript_detail,
                                    valueScriptId_title, valueScriptId_location, valueScriptId_description,
                                    valueScript_title, valueScript_location,
                                    valueScript_description, value_save_1).doAction()

        if actionName == CheckEditCalendarsAction.getActionName():
            value_task_1 = row[1].value
            value_edit = row[2].value
            value_id_title_2 = row[3].value
            value_title_2 = row[4].value
            value_save_2 = row[5].value
            value_id_task_2 = row[6].value
            value_select_delete = row[7].value
            value_delete_task = row[8].value
            CheckEditCalendarsAction(value_task_1, value_edit, value_id_title_2, value_title_2, value_save_2,
                                     value_id_task_2, value_select_delete, value_delete_task,).doAction()

        if actionName == CheckEditorsAction.getActionName():
            valueScript_editor = row[1].value
            valueScript_clear = row[2].value
            valueScript_add_content = row[3].value
            CheckEditorsAction(valueScript_editor, valueScript_clear,
                               valueScript_add_content).doAction()

        if actionName == CheckChartsAction.getActionName():
            valueScript_line = row[1].value
            valueScript_bar = row[2].value
            valueScript_pie = row[3].value
            valueScript_financial = row[4].value
            CheckChartsAction(valueScript_line, valueScript_bar, valueScript_pie, valueScript_financial).doAction()


# EXCUE PROGRAM
main()
